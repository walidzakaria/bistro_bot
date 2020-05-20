import openpyxl as xl
import os
from models import StaticVariables, ComparisonRecord

# Update Excel
app_dir = os.path.dirname(__file__)
file_path = os.path.join(app_dir, 'data_file.xlsx')
data_file = xl.load_workbook(file_path)
data_sheet = data_file.worksheets[0]


# data_sheet.cell(row=1, column=1).value = "test writing"
# print(data_sheet.max_row)


def get_last_record():
    StaticVariables.last_record = data_sheet.max_row


def save_record(comparison_record):
    data_sheet.cell(row=StaticVariables.current_record, column=8).value = comparison_record.fti_price
    data_sheet.cell(row=StaticVariables.current_record, column=9).value = comparison_record.competitor_name
    data_sheet.cell(row=StaticVariables.current_record, column=10).value = comparison_record.competitor_price


def save_work():
    data_file.save(file_path)


def get_record(row_index):
    result = ComparisonRecord(
        destination=data_sheet.cell(row=row_index, column=1).value,
        arrival=data_sheet.cell(row=row_index, column=2).value,
        departure=data_sheet.cell(row=row_index, column=3).value,
        duration_from=data_sheet.cell(row=row_index, column=4).value,
        duration_till=data_sheet.cell(row=row_index, column=5).value,
        hotel_name=data_sheet.cell(row=row_index, column=6).value,
        board=data_sheet.cell(row=row_index, column=7).value,
        fti_price=None,
        competitor_name=None,
        competitor_price=None
    )
    return result
